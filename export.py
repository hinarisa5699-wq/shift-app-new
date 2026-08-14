"""
export.py — Excel / CSV エクスポートモジュール
介護シフト自動作成アプリ

生成されたシフトデータを、整形済みの Excel ファイル (.xlsx) または
CSV ファイルとして出力する。
"""

import calendar
import csv
import io
import os
import re
from datetime import date
from io import BytesIO

import jpholiday
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet as WS   # PAPERSIZE_A3 / A4 定数用
from openpyxl.worksheet.properties import PageSetupProperties

# ---------------------------------------------------------------------------
# 定数: アサインメント → 日本語表示ラベル
# ---------------------------------------------------------------------------
ASSIGNMENT_LABELS = {
    "day_pattern1":    "デイ8:30-17:30",
    "day_pattern2":    "デイ9:00-16:00",
    "day_pattern3":    "デイ午前のみ",
    "day_pattern4":    "デイ午後のみ",
    "early":           "早番7:30-16:30",
    "late":            "遅番9:30-18:30",
    "nurse_short":     "看護9:30-13:30",
    "visit_am":        "訪問午前のみ",
    "visit_pm":        "訪問午後のみ",
    "day_p3_visit_pm": "兼務(デイ→訪問)",
    "visit_am_day_p4": "兼務(訪問→デイ)",
    "cooking_1":      "調理①6-8",
    "cooking_2":    "調理②8-13",
    "cooking_3":       "調理③12-19",
    "cooking_4":       "調理④6-13",
    "cooking_5":        "調理⑤9-15",
    # 旧名の後方互換
    "day_am":          "デイ午前のみ",
    "day_pm":          "デイ午後のみ",
    "day_am_visit_pm": "兼務(デイ→訪問)",
    "visit_am_day_pm": "兼務(訪問→デイ)",
}

# カテゴリごとの背景色 (アサインメントセル)
ASSIGNMENT_FILL = {
    "day_pattern1":    PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "day_pattern2":    PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid"),
    "day_pattern3":    PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"),
    "day_pattern4":    PatternFill(start_color="BAE6FD", end_color="BAE6FD", fill_type="solid"),
    "early":           PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "late":            PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid"),
    "nurse_short":     PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid"),
    "visit_am":        PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "visit_pm":        PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "day_p3_visit_pm": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
    "visit_am_day_p4": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
    "cooking_1":      PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "cooking_2":    PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid"),
    "cooking_3":       PatternFill(start_color="FCD34D", end_color="FCD34D", fill_type="solid"),
    "cooking_4":       PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid"),
    "cooking_5":        PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid"),
    # 旧名の後方互換
    "day_am":          PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"),
    "day_pm":          PatternFill(start_color="BAE6FD", end_color="BAE6FD", fill_type="solid"),
    "day_am_visit_pm": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
    "visit_am_day_pm": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
}

# 曜日名
WEEKDAY_NAMES = ["月", "火", "水", "木", "金", "土", "日"]

# 階別のデイ利用日／訪問日（曜日ルール）。d.weekday(): 0=月..6=日。
#   既定は従来の運用（3階 デイ=火金日/訪問=月木、2階 デイ=月木土/訪問=火金、外部デイ=水）。
#   実際の値は設定画面の階別営業曜日を configure_operating_days() で流し込む
#   （出力前に app.py が呼ぶ）。画面カレンダーと同じ内容が Excel/PDF にも出る。
_FLOOR3_DAY_SERVICE = {1, 4, 6}
_FLOOR3_VISIT = {0, 3}
_FLOOR2_DAY_SERVICE = {0, 3, 5}
_FLOOR2_VISIT = {1, 4}
_EXTERNAL_DAY_SERVICE = {2}


def configure_operating_days(
    floor3_day_service_days=None,
    floor3_visit_days=None,
    floor2_day_service_days=None,
    floor2_visit_days=None,
    external_day_service_days=None,
):
    """設定画面の階別営業曜日を反映する（0=月〜6=日）。

    None を渡した項目は既定値のまま据え置く。空リストは「その曜日は無し」として反映する。
    """
    global _FLOOR3_DAY_SERVICE, _FLOOR3_VISIT
    global _FLOOR2_DAY_SERVICE, _FLOOR2_VISIT, _EXTERNAL_DAY_SERVICE

    def _norm(days):
        return {int(x) for x in days if 0 <= int(x) <= 6}

    if floor3_day_service_days is not None:
        _FLOOR3_DAY_SERVICE = _norm(floor3_day_service_days)
    if floor3_visit_days is not None:
        _FLOOR3_VISIT = _norm(floor3_visit_days)
    if floor2_day_service_days is not None:
        _FLOOR2_DAY_SERVICE = _norm(floor2_day_service_days)
    if floor2_visit_days is not None:
        _FLOOR2_VISIT = _norm(floor2_visit_days)
    if external_day_service_days is not None:
        _EXTERNAL_DAY_SERVICE = _norm(external_day_service_days)


def _is_visit_weekday(d):
    """日付 d が訪問の営業日か（2階・3階いずれかが訪問の曜日）。"""
    return d.weekday() in (_FLOOR3_VISIT | _FLOOR2_VISIT)


def _floor_annotation(d):
    """日付 d の階別 デイ/訪問 注記（例 'デイ3階 訪2階'）。無ければ空文字。"""
    wd = d.weekday()
    parts = []
    if wd in _FLOOR3_DAY_SERVICE:
        parts.append("デイ3階")
    if wd in _FLOOR2_DAY_SERVICE:
        parts.append("デイ2階")
    if wd in _EXTERNAL_DAY_SERVICE:
        parts.append("外部デイ")
    if wd in _FLOOR3_VISIT:
        parts.append("訪3階")
    if wd in _FLOOR2_VISIT:
        parts.append("訪2階")
    return " ".join(parts)

# サマリー列ヘッダー (ケア)
# 値の並びは [day_am, day_pm, visit_am, visit_pm, ...]。ラベルは実体に一致させる。
SUMMARY_HEADERS = ["デイ午前", "デイ午後", "訪問午前", "訪問午後", "兼務者数", "オンコール"]

# サマリー列ヘッダー (調理)
COOK_SUMMARY_HEADERS = ["調理配置数"]

# ---------------------------------------------------------------------------
# スタイル定義
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="メイリオ", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="メイリオ", bold=True, size=16)
NORMAL_FONT = Font(name="メイリオ", size=10)
SATURDAY_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
SUNDAY_FILL = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
ALERT_FONT = Font(name="メイリオ", size=10, color="CC0000")
WARNING_HEADER_FILL = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
WARNING_HEADER_FONT = Font(name="メイリオ", bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# ⑧ 祝日行の背景色
HOLIDAY_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

# ③ 相談員事務スロットラベル
DESK_SLOT_LABELS = ["9-11時", "11-13時", "13-15時", "15-17時"]

# 兼務パターンは休憩はあるが相談業務なし
_NO_COUNSELOR_PATTERNS = {"day_p3_visit_pm", "visit_am_day_p4"}

# デイ午前に寄与するアサインメント
_DAY_AM_SET = {"day_pattern1", "day_pattern2", "day_pattern3", "day_p3_visit_pm",
               "day_am", "day_am_visit_pm", "late"}
# デイ午後に寄与するアサインメント
_DAY_PM_SET = {"day_pattern1", "day_pattern2", "day_pattern4", "visit_am_day_p4",
               "day_pm", "visit_am_day_pm", "early", "late"}
# 訪問午前
_VISIT_AM_SET = {"visit_am", "visit_am_day_p4", "visit_am_day_pm"}
# 訪問午後
_VISIT_PM_SET = {"visit_pm", "day_p3_visit_pm", "day_am_visit_pm"}
# 兼務
_DUAL_SET = {"day_p3_visit_pm", "visit_am_day_p4", "day_am_visit_pm", "visit_am_day_pm"}
# 調理
_COOK_SET = {"cooking_1", "cooking_2", "cooking_3", "cooking_4", "cooking_5"}
# 介護の「出勤」枠（デイ午前/午後・訪問・早番/遅番・看護短時間）
_CARE_WORK_SET = (
    _DAY_AM_SET | _DAY_PM_SET | _VISIT_AM_SET | _VISIT_PM_SET
    | {"early", "late", "nurse_short"}
)
_NURSE_PT_NAME_ALIASES = {"看護師", "PT", "理学療法士"}
_NURSE_PT_CODE_ALIASES = {"nurse", "pt"}


def _is_nurse_or_pt_staff(staff: dict) -> bool:
    """看護師/PTはコード優先で判定し、旧名称も受け入れる。"""
    qual_codes = {
        code for code in staff.get("qualification_codes", [])
        if isinstance(code, str) and code
    }
    if qual_codes.intersection(_NURSE_PT_CODE_ALIASES):
        return True

    qual_names = {
        name for name in staff.get("qualifications", [])
        if isinstance(name, str) and name
    }
    return bool(qual_names.intersection(_NURSE_PT_NAME_ALIASES))


# ---------------------------------------------------------------------------
# ヘルパー: 日ごとの配置データを集計する
# ---------------------------------------------------------------------------
def _build_daily_data(shifts_data, staff_list, year, month):
    """
    日付別・職員別の配置マップと、日付別サマリーを構築する。
    """
    num_days = calendar.monthrange(year, month)[1]
    dates = [date(year, month, d) for d in range(1, num_days + 1)]

    assignment_map = {}
    phone_duty_map = {}
    desk_slot_map = {}  # ③ {date_str: {staff_id: [slot_idx, ...]}}
    break_map = {}      # ① {date_str: {staff_id: "12:00"}}
    bath_map = {}       # お風呂当番 {date_str: {staff_id: "中"/"外"}}
    meal_map = {}       # 食事介助 {date_str: {staff_id: "12:00-13:00"}}
    for item in shifts_data:
        d_str = item["date"]
        sid = item["staff_id"]
        asgn = item.get("assignment", "")
        if d_str not in assignment_map:
            assignment_map[d_str] = {}
        assignment_map[d_str][sid] = asgn
        if item.get("is_phone_duty"):
            if d_str not in phone_duty_map:
                phone_duty_map[d_str] = []
            phone_duty_map[d_str].append(item.get("staff_name", f"ID:{sid}"))
        # ③ 相談員事務スロット
        slots = item.get("counselor_desk_slots")
        if slots:
            if d_str not in desk_slot_map:
                desk_slot_map[d_str] = {}
            desk_slot_map[d_str][sid] = slots
        # ① 休憩開始時刻
        bs = item.get("break_start")
        if bs:
            if d_str not in break_map:
                break_map[d_str] = {}
            break_map[d_str][sid] = bs
        # お風呂当番（中/外）
        bath = item.get("bath_role")
        if bath:
            bath_map.setdefault(d_str, {})[sid] = bath
        # 食事介助
        meal = item.get("meal_assist")
        if meal:
            meal_map.setdefault(d_str, {})[sid] = meal

    # ② 看護師/PTはデイ人数カウントから除外
    nurse_pt_ids = set()
    for st in staff_list:
        if _is_nurse_or_pt_staff(st):
            nurse_pt_ids.add(st["id"])

    summary_map = {}
    for d in dates:
        d_str = d.isoformat()
        day_assignments = assignment_map.get(d_str, {})

        day_am = 0
        day_pm = 0
        visit_am = 0
        visit_pm = 0
        dual = 0
        cook_total = 0

        for sid, asgn in day_assignments.items():
            is_nurse_pt = sid in nurse_pt_ids
            if asgn in _DAY_AM_SET and not is_nurse_pt:
                day_am += 1
            if asgn in _DAY_PM_SET and not is_nurse_pt:
                day_pm += 1
            if asgn in _VISIT_AM_SET:
                visit_am += 1
            elif asgn == "early" and _is_visit_weekday(d):
                # 訪問営業日の早番(7:30-16:30)は「午前訪問＋午後デイ」。
                #   シフト自動作成側も早番を午前訪問の担い手として数えているため、
                #   集計行でも訪問午前に算入する（従来は0名と表示され誤解の元だった）。
                visit_am += 1
            if asgn in _VISIT_PM_SET:
                visit_pm += 1
            if asgn in _DUAL_SET:
                dual += 1
            # 調理配置数: 種類マスタ化対応で cooking_* を前方一致で計数（新種類も含む）
            if asgn.startswith("cooking_") or asgn in _COOK_SET:
                cook_total += 1

        summary_map[d_str] = {
            "day_am": day_am,
            "day_pm": day_pm,
            "visit_am": visit_am,
            "visit_pm": visit_pm,
            "dual": dual,
            "cook_total": cook_total,
        }

    return dates, assignment_map, summary_map, phone_duty_map, desk_slot_map, break_map, bath_map, meal_map


# ---------------------------------------------------------------------------
# ヘルパー: 縦＝職員名・横＝日付 のシートを 1 枚書き込む
# ---------------------------------------------------------------------------
def _date_column_fill(d):
    """日付列の背景色（祝日 > 土 > 日）。平日は None。"""
    if jpholiday.is_holiday(d):
        return HOLIDAY_FILL
    weekday_idx = d.weekday()
    if weekday_idx == 5:
        return SATURDAY_FILL
    if weekday_idx == 6:
        return SUNDAY_FILL
    return None


def _staff_name_label(staff: dict, is_cook: bool) -> str:
    """職員名セルの表示（ケアは資格を併記）。"""
    if is_cook:
        return staff["name"]
    quals = staff.get("qualifications", [])
    if quals:
        return f"{staff['name']}\n({'/'.join(quals)})"
    return staff["name"]


def _care_cell_text(d_str, sid, assignment_map, bath_map, desk_slot_map):
    """ケアスタッフ 1 セルの (assignment, 表示テキスト) を組み立てる。"""
    asgn = assignment_map.get(d_str, {}).get(sid, "")
    text = ASSIGNMENT_LABELS.get(asgn, "")

    # お風呂当番（中/外）
    bath_role = bath_map.get(d_str, {}).get(sid)
    if bath_role:
        text += f"\n{bath_role}介助"

    # 休憩時間・食事介助ラベルは表示しない（要望により非表示）

    # ③ 相談員事務スロット
    desk_slots = desk_slot_map.get(d_str, {}).get(sid)
    if desk_slots:
        if set(desk_slots) >= {0, 1, 2, 3}:
            text += "\n相談（終日）"
        else:
            slot_texts = [DESK_SLOT_LABELS[si] for si in desk_slots if si < len(DESK_SLOT_LABELS)]
            if slot_texts:
                text += f"\n相談:{','.join(slot_texts)}"
    return asgn, text


def _build_parking_map(shifts_data):
    """依頼文24: shifts_data の parking_label から {date_iso: {staff_id: ラベル}} を作る。"""
    pm = {}
    for item in (shifts_data or []):
        lab = item.get("parking_label")
        if lab:
            pm.setdefault(item["date"], {})[item["staff_id"]] = lab
    return pm


def _parking_tag(label):
    """駐車枠ラベルの表示文字（枠番号は P付き、溢れは「コイン」）。"""
    if not label:
        return ""
    return "コイン" if label == "コイン" else f"P{label}"


def _append_parking(text, parking_map, d_str, sid):
    """セル本文に駐車場ラベルを改行で追記（車通勤・出勤者のみ map に存在）。"""
    tag = _parking_tag((parking_map or {}).get(d_str, {}).get(sid))
    if not tag:
        return text
    return f"{text}\n{tag}" if text else tag


# ---------------------------------------------------------------------------
# 印刷レイアウト（用紙・文字サイズ・行高）
#   介護=A3横 / 調理=A4横。どちらも「1ページに収めつつ枠いっぱいに大きく」。
#
#   Excel の「1ページに収める」は縮小しかしない（100%超に拡大はしない）ため、
#   印刷される文字の大きさは  フォントpt × 縮小率  で決まる。
#   縮小率は 内容の横幅 ÷ 印刷可能幅 で決まるので、用紙を大きくする(A3)か
#   フォントを上げるかのどちらかでしか実際の文字は大きくならない。
#   そこで縮小率を先に見積もり、印刷後に狙いのサイズになるようフォントを逆算する。
# ---------------------------------------------------------------------------
_MM_PER_PX = 25.4 / 96      # 列幅(px) → mm
_MM_PER_PT = 25.4 / 72      # 行高(pt) → mm
_PAPER_MM = {"A4": (297.0, 210.0), "A3": (420.0, 297.0)}   # いずれも横向き
_PRINT_MARGIN_IN = 0.2      # 上下左右の余白（インチ）。狭めて印刷領域を稼ぐ。

# 印刷後にこの pt 相当で出したい、という狙いの文字サイズ
_TARGET_PRINTED_PT = {"data": 11.0, "header": 11.5, "title": 16.0}
_FONT_PT_RANGE = (10.0, 26.0)   # 逆算したフォントサイズの下限・上限


def _col_width_to_mm(width_units):
    """Excel の列幅(文字数単位) を mm に換算する。"""
    return (width_units * 7 + 5) * _MM_PER_PX


def _plan_print_layout(is_cook, num_days, name_width, date_width, total_width,
                       n_data_rows, n_summary_rows, fit_one_page):
    """用紙・フォントサイズ・行高をまとめて決める。

    戻り値: dict(paper, paper_code, data_pt, header_pt, title_pt,
                 data_row_h, header_row_h, title_row_h, dow_row_h)
    """
    paper = "A4" if is_cook else "A3"
    paper_w, paper_h = _PAPER_MM[paper]
    margin_mm = _PRINT_MARGIN_IN * 25.4
    usable_w = paper_w - margin_mm * 2
    usable_h = paper_h - margin_mm * 2

    # --- 横方向: 内容幅から縮小率を見積もる ---
    content_w = (
        _col_width_to_mm(name_width)
        + _col_width_to_mm(date_width) * num_days
        + _col_width_to_mm(total_width)
    )
    if fit_one_page and content_w > 0:
        # 1ページに収める指定のときだけ縮小がかかる（拡大はされないので上限1.0）
        scale = min(1.0, usable_w / content_w)
    else:
        scale = 1.0   # 横は複数ページに流すので等倍で印刷される

    def _solve_pt(kind):
        lo, hi = _FONT_PT_RANGE
        return round(min(hi, max(lo, _TARGET_PRINTED_PT[kind] / scale)), 1)

    data_pt = _solve_pt("data")
    header_pt = _solve_pt("header")
    title_pt = _solve_pt("title")

    # サマリー行のラベル（「調理配置数」「オンコール」等・全角5文字＝10単位）は
    # 折り返さずに職員名列へ収めたい。列幅で頭打ちにする（列幅の単位は既定フォント11pt基準）。
    #   +0.5 は丸め・字間の安全代（ぴったりだと1文字だけ溢れることがある）。
    _SUMMARY_LABEL_UNITS = 10
    summary_label_pt = round(min(data_pt, name_width * 11.0 / (_SUMMARY_LABEL_UNITS + 0.5)), 1)

    # --- 縦方向: 残り高さをデータ行に配って「枠いっぱい」にする ---
    title_row_h = round(title_pt * 1.5, 1)
    header_row_h = round(header_pt * 1.6, 1)
    # 曜日行は「曜日／祝日名／デイ・訪問注記」で最大3行になる（例: 火・山の日・訪問）
    dow_row_h = round(header_pt * 1.25 * 3 + 2, 1)
    summary_row_h = round(data_pt * 1.6, 1)

    fixed_h_pt = title_row_h + header_row_h + dow_row_h + summary_row_h * n_summary_rows
    # 縮小率は横で決まるので、縦もその縮小率で刷られる前提で使える高さを逆算する。
    # 0.95 は安全マージン: 縦がわずかでも溢れると fitToHeight が効いて更に縮小され、
    # 狙ったフォントサイズより小さく印刷されてしまうため。
    avail_pt = (usable_h / _MM_PER_PT) / scale * 0.95 - fixed_h_pt
    min_row_h = round(data_pt * 1.9, 1)     # 2行ぶん折り返しても潰れない最低限
    if n_data_rows > 0:
        # 上限90pt: 職員が少ない月に1行が間延びしすぎないように抑える
        data_row_h = max(min_row_h, min(90.0, avail_pt / n_data_rows))
    else:
        data_row_h = min_row_h

    return {
        "paper": paper,
        "paper_code": WS.PAPERSIZE_A3 if paper == "A3" else WS.PAPERSIZE_A4,
        "data_pt": data_pt,
        "header_pt": header_pt,
        "title_pt": title_pt,
        "summary_label_pt": summary_label_pt,
        "data_row_h": round(data_row_h, 1),
        "header_row_h": header_row_h,
        "title_row_h": title_row_h,
        "dow_row_h": dow_row_h,
        "summary_row_h": summary_row_h,
    }


def _write_group_sheet(
    ws,
    group_staff,
    dates,
    year,
    month,
    *,
    assignment_map,
    summary_map,
    phone_duty_map,
    desk_slot_map,
    bath_map,
    warnings_data,
    is_cook,
    fit_one_page=False,
    parking_map=None,
    title_override=None,
):
    """1 グループ（介護 or 調理）を「縦＝職員名・横＝日付」で 1 シートに書き込む。
    fit_one_page=True のとき、印刷時に横もA4横1ページに収める（依頼文25・4分割向け）。
    title_override を渡すとタイトル文字列を差し替える（PDFと書式を揃える用）。"""
    num_days = len(dates)
    name_col = 1
    first_date_col = 2
    last_date_col = first_date_col + num_days - 1
    total_col = last_date_col + 1   # 出勤日数列

    title_label = "調理スタッフ" if is_cook else "介護スタッフ"
    off_token = "cook_off" if is_cook else "off"

    # --- 印刷レイアウト（用紙・文字サイズ・行高）を先に決める ---
    #   ここで決めたサイズをこのシート内の全セルに使う（モジュール定数のフォントは
    #   他のシートでも使い回されるため、ここでは触らずローカルに作る）。
    name_width = 18 if not is_cook else 12.875
    date_width = 14 if is_cook else 13   # 調理は時間ラベルが長い
    total_width = 7
    n_summary_rows = 1 if is_cook else 6
    layout = _plan_print_layout(
        is_cook, num_days, name_width, date_width, total_width,
        n_data_rows=len(group_staff), n_summary_rows=n_summary_rows,
        fit_one_page=fit_one_page,
    )

    title_font = Font(name="メイリオ", bold=True, size=layout["title_pt"])
    header_font_wrap = Font(name="メイリオ", bold=True, color="FFFFFF", size=layout["header_pt"])
    label_font = Font(name="メイリオ", bold=True, size=layout["data_pt"])
    body_font = Font(name="メイリオ", size=layout["data_pt"])
    summary_label_font = Font(name="メイリオ", bold=True, size=layout["summary_label_pt"])
    alert_body_font = Font(name="メイリオ", size=layout["data_pt"], color="CC0000")

    # --- タイトル行 ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_col)
    title_value = title_override or f"{year}年{month}月 シフト表（{title_label}）"
    title_cell = ws.cell(row=1, column=1, value=title_value)
    title_cell.font = title_font
    title_cell.alignment = CENTER_ALIGN

    header_row1 = 2   # 日付（M/D）
    header_row2 = 3   # 曜日
    data_start_row = 4

    # --- 職員名ヘッダー（2 行ぶち抜き）---
    ws.merge_cells(start_row=header_row1, start_column=name_col, end_row=header_row2, end_column=name_col)
    c = ws.cell(row=header_row1, column=name_col, value="職員名")
    c.font = header_font_wrap
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN_BORDER
    ws.cell(row=header_row2, column=name_col).border = THIN_BORDER

    # --- 出勤日数ヘッダー（2 行ぶち抜き）---
    ws.merge_cells(start_row=header_row1, start_column=total_col, end_row=header_row2, end_column=total_col)
    c = ws.cell(row=header_row1, column=total_col, value="出勤\n日数")
    c.font = header_font_wrap
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN_BORDER
    ws.cell(row=header_row2, column=total_col).border = THIN_BORDER

    # --- 日付ヘッダー（1 行目＝M/D、2 行目＝曜日 / 祝日名）---
    for i, d in enumerate(dates):
        col = first_date_col + i
        col_fill = _date_column_fill(d)

        c1 = ws.cell(row=header_row1, column=col, value=f"{d.month}/{d.day}")
        c1.font = header_font_wrap
        c1.fill = HEADER_FILL
        c1.alignment = CENTER_ALIGN
        c1.border = THIN_BORDER

        weekday_name = WEEKDAY_NAMES[d.weekday()]
        if jpholiday.is_holiday(d):
            dow_value = f"{weekday_name}\n{jpholiday.is_holiday_name(d)}"
        else:
            dow_value = weekday_name
        # 階別 デイ/訪問 注記（画面カレンダーと同じ）を曜日の下に追記
        _floor = _floor_annotation(d)
        if _floor:
            dow_value = f"{dow_value}\n{_floor}"
        c2 = ws.cell(row=header_row2, column=col, value=dow_value)
        c2.font = body_font
        c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c2.border = THIN_BORDER
        if col_fill:
            c2.fill = col_fill

    # --- 職員ごとのデータ行 ---
    for r_off, s in enumerate(group_staff):
        row = data_start_row + r_off
        sid = s["id"]

        name_cell = ws.cell(row=row, column=name_col, value=_staff_name_label(s, is_cook))
        name_cell.font = body_font
        name_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        name_cell.border = THIN_BORDER

        work_days = 0
        for i, d in enumerate(dates):
            col = first_date_col + i
            d_str = d.isoformat()
            col_fill = _date_column_fill(d)

            if is_cook:
                asgn = assignment_map.get(d_str, {}).get(sid, "")
                text = ASSIGNMENT_LABELS.get(asgn, "")
            else:
                asgn, text = _care_cell_text(d_str, sid, assignment_map, bath_map, desk_slot_map)

            # 駐車場（依頼文24）: 車通勤・出勤者のセルに枠/コインを追記
            text = _append_parking(text, parking_map, d_str, sid)

            cell = ws.cell(row=row, column=col, value=text)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

            # 調理はPDFと書式を揃えるため、種類別の色付け（オレンジ系）はせず
            # 週末・祝日の列色のみを敷く（PDFと同じ白基調）。介護は従来通り色分け。
            if (not is_cook) and asgn in ASSIGNMENT_FILL:
                cell.fill = ASSIGNMENT_FILL[asgn]
            elif col_fill:
                cell.fill = col_fill

            if asgn not in (off_token, ""):
                work_days += 1

        total_cell = ws.cell(row=row, column=total_col, value=work_days)
        total_cell.font = label_font
        total_cell.alignment = CENTER_ALIGN
        total_cell.border = THIN_BORDER

    # --- サマリー行（日付ごとの配置数。横＝日付に合わせて行として並べる）---
    summary_start_row = data_start_row + len(group_staff)
    if is_cook:
        summary_rows = [("調理配置数", "cook_total", "understaffed_cook")]
    else:
        summary_rows = [
            ("デイ午前", "day_am", "understaffed_day_am"),
            ("デイ午後", "day_pm", "understaffed_day_pm"),
            ("訪問午前", "visit_am", "understaffed_visit_am"),
            ("訪問午後", "visit_pm", "understaffed_visit_pm"),
            ("兼務者数", "dual", "dual_shortage"),
            ("オンコール", "_phone", None),
        ]

    for r_off, (label, key, warn_type) in enumerate(summary_rows):
        row = summary_start_row + r_off
        label_cell = ws.cell(row=row, column=name_col, value=label)
        label_cell.font = summary_label_font
        label_cell.fill = SATURDAY_FILL
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = THIN_BORDER

        for i, d in enumerate(dates):
            col = first_date_col + i
            d_str = d.isoformat()
            summary = summary_map.get(d_str, {})

            if key == "_phone":
                names = phone_duty_map.get(d_str, [])
                val = ", ".join(names) if names else ""
            else:
                val = summary.get(key, 0)

            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

            is_alert = False
            if warn_type:
                for w in warnings_data:
                    if w.get("date") != d_str:
                        continue
                    wt = w.get("warning_type", "")
                    if warn_type == "understaffed_cook":
                        if wt.startswith("understaffed_cook"):
                            is_alert = True
                            break
                    elif wt == warn_type:
                        is_alert = True
                        break
            if is_alert:
                cell.fill = ALERT_FILL
                cell.font = alert_body_font

        ws.cell(row=row, column=total_col, value="").border = THIN_BORDER

    # --- 列幅・行高・固定・印刷設定 ---
    #   列幅は _plan_print_layout の縮小率計算と同じ値を使う（ズレると狙いが外れる）。
    ws.column_dimensions[get_column_letter(name_col)].width = name_width   # 職員名・資格
    for i in range(num_days):
        ws.column_dimensions[get_column_letter(first_date_col + i)].width = date_width
    ws.column_dimensions[get_column_letter(total_col)].width = total_width

    # 行高: 残った縦の余白をデータ行に配って枠いっぱいに使う
    ws.row_dimensions[1].height = layout["title_row_h"]
    ws.row_dimensions[header_row1].height = layout["header_row_h"]
    ws.row_dimensions[header_row2].height = layout["dow_row_h"]   # 曜日＋祝日名＋注記
    for r_off in range(len(group_staff)):
        ws.row_dimensions[data_start_row + r_off].height = layout["data_row_h"]
    for r_off in range(n_summary_rows):
        ws.row_dimensions[summary_start_row + r_off].height = layout["summary_row_h"]

    # 用紙: 介護=A3横 / 調理=A4横（どちらも枠いっぱい・1ページに収める）
    ws.page_setup.paperSize = layout["paper_code"]
    ws.page_setup.orientation = "landscape"
    # 依頼文25: 4分割（前半15日/後半16日）は横も1ページに収める
    ws.page_setup.fitToWidth = 1 if fit_one_page else 0   # 0=日付ぶん複数ページ
    ws.page_setup.fitToHeight = 1   # 縦（職員）は 1 ページに収める
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=_PRINT_MARGIN_IN, right=_PRINT_MARGIN_IN,
        top=_PRINT_MARGIN_IN, bottom=_PRINT_MARGIN_IN,
        header=0.1, footer=0.1,
    )
    # 余白ぶんだけ左右に寄らないよう中央に置く
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.print_title_rows = "1:3"      # タイトル＋日付見出しを各ページ先頭に繰り返す
    ws.print_title_cols = "A:A"      # 職員名列を各ページ左端に繰り返す
    ws.freeze_panes = ws.cell(row=data_start_row, column=first_date_col).coordinate


# ---------------------------------------------------------------------------
# Excel エクスポート
# ---------------------------------------------------------------------------
def _register_cook_labels(cook_labels):
    """調理シフト種類マスタのラベルを取り込む（既存①〜⑤のラベルは保持＝setdefault）。
    新しく追加された種類(cooking_6…)のラベル表示に使う。"""
    if not cook_labels:
        return
    for code, label in cook_labels.items():
        if label:
            ASSIGNMENT_LABELS.setdefault(code, label)


def export_excel(
    shifts_data: list,
    warnings_data: list,
    staff_list: list,
    year: int,
    month: int,
    oncall_map: dict = None,
    cook_labels: dict = None,
) -> BytesIO:
    """Excel 形式でシフト表を出力する。

    レイアウトは「縦＝職員名・横＝日付」。介護スタッフと調理スタッフを
    別シート（1 枚目＝介護、2 枚目＝調理）に分けて出力する。
    """
    _register_cook_labels(cook_labels)
    wb = Workbook()

    dates, assignment_map, summary_map, phone_duty_map, desk_slot_map, break_map, bath_map, meal_map = _build_daily_data(
        shifts_data, staff_list, year, month
    )
    # オンコール（電話当番）は出勤と独立した別データで上書き
    if oncall_map is not None:
        phone_duty_map = {d: [n] for d, n in oncall_map.items() if n}

    parking_map = _build_parking_map(shifts_data)
    care_staff = [s for s in staff_list if s.get("department") != "cooking"]
    cook_staff = [s for s in staff_list if s.get("department") == "cooking"]

    # --- 1 枚目: 介護スタッフ ---
    ws_care = wb.active
    ws_care.title = "介護スタッフ"
    _write_group_sheet(
        ws_care, care_staff, dates, year, month,
        assignment_map=assignment_map,
        summary_map=summary_map,
        phone_duty_map=phone_duty_map,
        desk_slot_map=desk_slot_map,
        bath_map=bath_map,
        warnings_data=warnings_data,
        is_cook=False,
        parking_map=parking_map,
    )

    # --- 2 枚目: 調理スタッフ ---
    if cook_staff:
        ws_cook = wb.create_sheet(title="調理スタッフ")
        _write_group_sheet(
            ws_cook, cook_staff, dates, year, month,
            assignment_map=assignment_map,
            summary_map=summary_map,
            phone_duty_map=phone_duty_map,
            desk_slot_map=desk_slot_map,
            bath_map=bath_map,
            warnings_data=warnings_data,
            is_cook=True,
            parking_map=parking_map,
        )

    # --- 警告シート ---
    if warnings_data:
        ws_warn = wb.create_sheet(title="警告一覧")

        warn_headers = ["日付", "種別", "内容"]
        for col_idx, header in enumerate(warn_headers, start=1):
            cell = ws_warn.cell(row=1, column=col_idx, value=header)
            cell.font = WARNING_HEADER_FONT
            cell.fill = WARNING_HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        for row_offset, warn in enumerate(warnings_data):
            row = 2 + row_offset
            ws_warn.cell(row=row, column=1, value=warn.get("date", "")).font = NORMAL_FONT
            ws_warn.cell(row=row, column=1).border = THIN_BORDER
            ws_warn.cell(row=row, column=1).alignment = CENTER_ALIGN

            ws_warn.cell(
                row=row, column=2, value=warn.get("warning_type", "")
            ).font = NORMAL_FONT
            ws_warn.cell(row=row, column=2).border = THIN_BORDER
            ws_warn.cell(row=row, column=2).alignment = CENTER_ALIGN

            ws_warn.cell(
                row=row, column=3, value=warn.get("message", "")
            ).font = NORMAL_FONT
            ws_warn.cell(row=row, column=3).border = THIN_BORDER

        ws_warn.column_dimensions["A"].width = 12
        ws_warn.column_dimensions["B"].width = 18
        ws_warn.column_dimensions["C"].width = 50

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# CSV エクスポート
# ---------------------------------------------------------------------------
def export_csv(
    shifts_data: list,
    warnings_data: list,
    staff_list: list,
    year: int,
    month: int,
    oncall_map: dict = None,
    cook_labels: dict = None,
) -> str:
    """CSV 形式でシフト表を出力する。

    レイアウトは Excel と同じ「縦＝職員名・横＝日付」。1 ファイル内に
    介護スタッフ・調理スタッフのブロックを順に出力する（CSV はシートを
    持てないため、空行で区切る）。
    """
    _register_cook_labels(cook_labels)
    dates, assignment_map, summary_map, phone_duty_map, desk_slot_map, break_map, bath_map, meal_map = _build_daily_data(
        shifts_data, staff_list, year, month
    )
    # オンコール（電話当番）は出勤と独立した別データで上書き
    if oncall_map is not None:
        phone_duty_map = {d: [n] for d, n in oncall_map.items() if n}

    care_staff = [s for s in staff_list if s.get("department") != "cooking"]
    cook_staff = [s for s in staff_list if s.get("department") == "cooking"]

    output = io.StringIO()
    writer = csv.writer(output)

    # 日付ヘッダー（M/D(曜)）。各ブロックは 職員名列＋日付列＋出勤日数列。
    date_headers = []
    for d in dates:
        date_headers.append(f"{d.month}/{d.day}({WEEKDAY_NAMES[d.weekday()]})")

    def _care_cell(d_str, sid):
        asgn = assignment_map.get(d_str, {}).get(sid, "")
        label = ASSIGNMENT_LABELS.get(asgn, "")
        parts = [label] if label else []
        bath_role = bath_map.get(d_str, {}).get(sid)
        if bath_role:
            parts.append(f"{bath_role}介助")
        # 休憩時間・食事介助の時刻は表示しない（要望により非表示。Excel・画面と統一）
        desk_slots = desk_slot_map.get(d_str, {}).get(sid)
        if desk_slots:
            parts.append("相談（終日）" if set(desk_slots) >= {0, 1, 2, 3}
                         else "相談:" + ",".join(
                             DESK_SLOT_LABELS[si] for si in desk_slots if si < len(DESK_SLOT_LABELS)))
        return asgn, " ".join(parts)

    def _write_block(title, group_staff, is_cook):
        off_token = "cook_off" if is_cook else "off"
        writer.writerow([title])
        writer.writerow(["職員名"] + date_headers + ["出勤日数"])

        for s in group_staff:
            sid = s["id"]
            name = s["name"]
            quals = s.get("qualifications", [])
            if not is_cook and quals:
                name = f"{name}({'/'.join(quals)})"

            cells = []
            work_days = 0
            for d in dates:
                d_str = d.isoformat()
                if is_cook:
                    asgn = assignment_map.get(d_str, {}).get(sid, "")
                    cells.append(ASSIGNMENT_LABELS.get(asgn, ""))
                else:
                    asgn, text = _care_cell(d_str, sid)
                    cells.append(text)
                if asgn not in (off_token, ""):
                    work_days += 1
            writer.writerow([name] + cells + [work_days])

        # サマリー行（日付ごとの配置数）
        if is_cook:
            summary_rows = [("調理配置数", "cook_total")]
        else:
            summary_rows = [
                ("デイ午前", "day_am"),
                ("デイ午後", "day_pm"),
                ("訪問午前", "visit_am"),
                ("訪問午後", "visit_pm"),
                ("兼務者数", "dual"),
                ("オンコール", "_phone"),
            ]
        for label, key in summary_rows:
            cells = []
            for d in dates:
                d_str = d.isoformat()
                if key == "_phone":
                    names = phone_duty_map.get(d_str, [])
                    cells.append(", ".join(names) if names else "")
                else:
                    cells.append(summary_map.get(d_str, {}).get(key, 0))
            writer.writerow([label] + cells + [""])

    _write_block("【介護スタッフ】", care_staff, is_cook=False)
    if cook_staff:
        writer.writerow([])
        _write_block("【調理スタッフ】", cook_staff, is_cook=True)

    csv_string = "\ufeff" + output.getvalue()
    return csv_string


# ---------------------------------------------------------------------------
# PDF エクスポート（表示専用。Excel/CSV と同じ集計を再利用）
# ---------------------------------------------------------------------------
# 同梱 CJK フォント（システムフォントに依存しない＝Renderでも文字化けしない）
_FONT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
_PDF_FONT_PATH = os.path.join(_FONT_DIR, "ipaexg.ttf")
_PDF_FONT = "ipaexg"

# PDF 色（RGB）
_PDF_HEADER_BG = (68, 114, 196)
_PDF_SAT_BG = (232, 240, 254)
_PDF_SUN_BG = (253, 232, 232)
_PDF_HOL_BG = (255, 243, 224)
_PDF_SUMMARY_BG = (238, 242, 255)
_PDF_ALERT_BG = (254, 226, 226)

# サマリー行: ラベル → (summary_mapキー or 特殊, 警告種別)
_PDF_CARE_SUMMARY = [
    ("デイ午前", "day_am", "understaffed_day_am"),
    ("デイ午後", "day_pm", "understaffed_day_pm"),
    ("訪問午前", "visit_am", "understaffed_visit_am"),
    ("訪問午後", "visit_pm", "understaffed_visit_pm"),
    ("兼務者数", "dual", "dual_shortage"),
    ("オンコール", "_phone", None),
]
_PDF_COOK_SUMMARY = [("調理配置数", "cook_total", "understaffed_cook")]


def _pdf_weekend_color(d):
    """日付列の背景色（祝日 > 土 > 日）。平日は None。"""
    if jpholiday.is_holiday(d):
        return _PDF_HOL_BG
    wd = d.weekday()
    if wd == 5:
        return _PDF_SAT_BG
    if wd == 6:
        return _PDF_SUN_BG
    return None


def _pdf_wrap(pdf, text, max_w, fs):
    """セル幅 max_w(mm) に収まるよう改行（CJKは空白が無いので文字単位で折返し）。"""
    pdf.set_font(_PDF_FONT, "", fs)
    lines = []
    for raw in str(text).split("\n"):
        if raw == "":
            continue
        cur = ""
        for ch in raw:
            if pdf.get_string_width(cur + ch) <= max_w - 1.2:
                cur += ch
            else:
                if cur:
                    lines.append(cur)
                cur = ch
        if cur:
            lines.append(cur)
    return lines or [""]


def _pdf_line_h(fs):
    """フォントサイズ(pt)に対する1行の高さ(mm)。"""
    return fs * 0.3528 * 1.18


def _pick_body_fs(pdf, staff_rows, summary_rows, name_w, date_w, row_h,
                  fs_max=22.0, fs_min=4.0, step=0.5):
    """依頼文25: 1行高 row_h に収まる範囲で最大の本文フォントサイズ(pt)を選ぶ。
    行数が少ない（=row_h が大きい）シートほど大きい値が返る。"""
    def fits(fs):
        max_lines = 1
        for name, cells, _w in staff_rows:
            max_lines = max(max_lines, len(_pdf_wrap(pdf, name, name_w, fs)))
            for c in cells:
                max_lines = max(max_lines, len(_pdf_wrap(pdf, c, date_w, fs)))
        for label, vals, _al in summary_rows:
            max_lines = max(max_lines, len(_pdf_wrap(pdf, label, name_w, fs)))
            for v in vals:
                max_lines = max(max_lines, len(_pdf_wrap(pdf, v, date_w, fs)))
        return max_lines * _pdf_line_h(fs) <= row_h - 0.6

    fs = fs_max
    while fs >= fs_min:
        if fits(fs):
            return fs
        fs -= step
    return fs_min


def _render_pdf_table(title, sel_dates, staff_rows, summary_rows, paper="A4"):
    """共通PDF描画（横・1ページ）。データ元は DB でも Excel でもよい。
    paper:        "A3"（介護）/ "A4"（調理）。Excel出力と同じ用紙に揃える。
    title:        タイトル文字列
    sel_dates:    列に対応する date のリスト（曜日/祝日色・M/D見出し用）
    staff_rows:   [(name, [cell_text, ...], work_str), ...]
    summary_rows: [(label, [val_str, ...], [alert_bool, ...]), ...]
    """
    from fpdf import FPDF  # 遅延 import

    # 依頼文25: 指定用紙の横1枚ちょうどに収めつつ、ページいっぱいに最大の文字で描画する。
    # 余白を最小化し、行高は行数に応じて usable_h を等分（行数の少ない調理は行が高く＝文字が大きくなる）、
    # 文字サイズは「1枚に収まる範囲で最大」を自動選択する。
    MARGIN = 4.0          # 余白（mm）最小限
    TITLE_H = 7.0         # タイトル帯の高さ
    GAP = 1.5             # タイトルと表の隙間

    pdf = FPDF(orientation="L", unit="mm", format=paper if paper in ("A3", "A4") else "A4")
    pdf.set_auto_page_break(False)
    pdf.add_font(_PDF_FONT, "", _PDF_FONT_PATH)
    pdf.set_margins(MARGIN, MARGIN, MARGIN)
    pdf.add_page()
    page_w = pdf.w
    page_h = pdf.h

    table_top = MARGIN + TITLE_H + GAP
    usable_w = page_w - MARGIN * 2
    usable_h = page_h - MARGIN - table_top
    n_dates = len(sel_dates)
    name_w = 30.0
    total_w = 13.0
    date_w = (usable_w - name_w - total_w) / max(n_dates, 1)
    total_rows = 3 + len(staff_rows) + len(summary_rows)  # ヘッダー3行ぶん（日付/曜日/階別注記）
    row_h = usable_h / max(total_rows, 1)

    def line_h(fs):
        return _pdf_line_h(fs)

    # 上限を大きく取り、収まる最大サイズを選ぶ（調理など行数が少ないシートは大きくなる）
    body_fs = _pick_body_fs(pdf, staff_rows, summary_rows, name_w, date_w, row_h)
    header_fs = min(body_fs + 0.5, 22.0)
    max_lines_per_cell = max(1, int((row_h - 0.6) / line_h(body_fs)))

    # タイトル（表幅に合わせ、フォントは表より少し大きく）
    title_fs = min(max(body_fs, 11.0), 16.0)
    pdf.set_font(_PDF_FONT, "", title_fs)
    pdf.set_xy(MARGIN, MARGIN)
    pdf.cell(usable_w, TITLE_H, title, align="C")

    def draw_cell(x, y, w, h, text, fs, *, fill=None, bold_color=None, align="C"):
        if fill:
            pdf.set_fill_color(*fill)
            pdf.rect(x, y, w, h, style="DF")
        else:
            pdf.rect(x, y, w, h, style="D")
        if text == "" or text is None:
            return
        lines = _pdf_wrap(pdf, text, w, fs)[:max_lines_per_cell]
        pdf.set_font(_PDF_FONT, "", fs)
        pdf.set_text_color(*(bold_color or (0, 0, 0)))
        lh = line_h(fs)
        ty = y + (h - lh * len(lines)) / 2
        for ln in lines:
            pdf.set_xy(x, ty)
            pdf.cell(w, lh, ln, align=align)
            ty += lh
        pdf.set_text_color(0, 0, 0)

    # --- ヘッダー行（日付 / 曜日＋階別注記）---
    #   曜日行を2段分の高さにして「曜日＋デイ○階/訪○階/外部デイ」を表示、
    #   データ行はさらに1段下げる（ヘッダ計3段）。
    y = table_top
    _floor_fs = max(header_fs - 3.0, 6.0)  # 階別注記は小さめフォント
    header_h = row_h * 3
    draw_cell(MARGIN, y, name_w, header_h, "職員名", header_fs,
              fill=_PDF_HEADER_BG, bold_color=(255, 255, 255))
    x = MARGIN + name_w
    for d in sel_dates:
        draw_cell(x, y, date_w, row_h, f"{d.month}/{d.day}", header_fs,
                  fill=_PDF_HEADER_BG, bold_color=(255, 255, 255))
        wd = WEEKDAY_NAMES[d.weekday()]
        if jpholiday.is_holiday(d):
            wd = f"{wd}/祝"
        draw_cell(x, y + row_h, date_w, row_h, wd, header_fs, fill=_pdf_weekend_color(d))
        # 階別注記（デイ○階／訪○階／外部デイ）を曜日の下に小さく
        _floor = _floor_annotation(d)
        draw_cell(x, y + row_h * 2, date_w, row_h, _floor, _floor_fs,
                  fill=_pdf_weekend_color(d))
        x += date_w
    draw_cell(x, y, total_w, header_h, "出勤\n日数", header_fs,
              fill=_PDF_HEADER_BG, bold_color=(255, 255, 255))

    # --- 職員行 ---（ヘッダは3段：日付/曜日/階別注記）
    y = table_top + row_h * 3
    for name, cells, work in staff_rows:
        draw_cell(MARGIN, y, name_w, row_h, name, body_fs, align="L")
        x = MARGIN + name_w
        for i, d in enumerate(sel_dates):
            draw_cell(x, y, date_w, row_h, cells[i] if i < len(cells) else "", body_fs,
                      fill=_pdf_weekend_color(d))
            x += date_w
        draw_cell(x, y, total_w, row_h, str(work), body_fs)
        y += row_h

    # --- サマリー行 ---
    for label, vals, alerts in summary_rows:
        draw_cell(MARGIN, y, name_w, row_h, label, body_fs, fill=_PDF_SUMMARY_BG, align="L")
        x = MARGIN + name_w
        for i, d in enumerate(sel_dates):
            alert = alerts[i] if i < len(alerts) else False
            draw_cell(x, y, date_w, row_h, vals[i] if i < len(vals) else "", body_fs,
                      fill=_PDF_ALERT_BG if alert else _pdf_weekend_color(d),
                      bold_color=(204, 0, 0) if alert else None)
            x += date_w
        draw_cell(x, y, total_w, row_h, "", body_fs)
        y += row_h

    buf = BytesIO()
    buf.write(bytes(pdf.output()))
    buf.seek(0)
    return buf


def _pdf_group_meta(group):
    is_cook = (group == "cooking")
    return is_cook, ("調理" if is_cook else "介護・看護"), (_PDF_COOK_SUMMARY if is_cook else _PDF_CARE_SUMMARY)


def _half_dates(dates, half):
    if half == "second":
        return [d for d in dates if d.day >= 16], "後半"
    return [d for d in dates if d.day <= 15], "前半"


def export_pdf(
    shifts_data: list,
    warnings_data: list,
    staff_list: list,
    year: int,
    month: int,
    group: str = "care",
    half: str = "first",
    oncall_map: dict = None,
    cook_labels: dict = None,
) -> BytesIO:
    """DB由来データから PDF を出力する（A4横・1ページ・表示専用）。
    group: "care"=介護・看護 / "cooking"=調理 / half: "first"=1〜15 / "second"=16〜末
    """
    _register_cook_labels(cook_labels)
    dates, assignment_map, summary_map, phone_duty_map, desk_slot_map, break_map, bath_map, meal_map = _build_daily_data(
        shifts_data, staff_list, year, month
    )
    if oncall_map is not None:
        phone_duty_map = {d: [n] for d, n in oncall_map.items() if n}

    is_cook, group_label, summary_defs = _pdf_group_meta(group)
    gstaff = [s for s in staff_list if (s.get("department") == "cooking") == is_cook]
    sel, half_label = _half_dates(dates, half)
    off_token = "cook_off" if is_cook else "off"
    parking_map = _build_parking_map(shifts_data)

    def cell_text(sid, d):
        """(表示テキスト, assignmentコード) を返す。care/cook で順序を統一。"""
        d_str = d.isoformat()
        if is_cook:
            asgn = assignment_map.get(d_str, {}).get(sid, "")
            text = ASSIGNMENT_LABELS.get(asgn, "")
        else:
            asgn, text = _care_cell_text(d_str, sid, assignment_map, bath_map, desk_slot_map)
        text = _append_parking(text, parking_map, d_str, sid)
        return text, asgn

    staff_rows = []
    for s in gstaff:
        sid = s["id"]
        cells, work = [], 0
        for d in sel:
            txt, asgn = cell_text(sid, d)
            cells.append(txt)
            if asgn not in (off_token, ""):
                work += 1
        staff_rows.append((_staff_name_label(s, is_cook), cells, str(work)))

    def has_warn(d_str, wtype):
        if not wtype:
            return False
        for w in warnings_data:
            if w.get("date") != d_str:
                continue
            wt = w.get("warning_type", "")
            if wtype == "understaffed_cook":
                if wt.startswith("understaffed_cook"):
                    return True
            elif wt == wtype:
                return True
        return False

    summary_rows = []
    for label, key, wtype in summary_defs:
        vals, alerts = [], []
        for d in sel:
            d_str = d.isoformat()
            if key == "_phone":
                names = phone_duty_map.get(d_str, [])
                vals.append(", ".join(names) if names else "")
            else:
                vals.append(str(summary_map.get(d_str, {}).get(key, 0)))
            alerts.append(has_warn(d_str, wtype))
        summary_rows.append((label, vals, alerts))

    first_d = sel[0].day if sel else 1
    last_d = sel[-1].day if sel else 1
    title = f"{year}年{month}月 シフト表（{group_label} {half_label}：{first_d}〜{last_d}日）"
    # 介護=A3横 / 調理=A4横（Excel出力と同じ用紙）
    return _render_pdf_table(title, sel, staff_rows, summary_rows,
                             paper="A4" if group == "cooking" else "A3")


# ---------------------------------------------------------------------------
# 職員1人ずつの月間カレンダーPDF（縦A4・1人1ページ）
# ---------------------------------------------------------------------------
# カレンダーは日曜始まり（列: 日 月 火 水 木 金 土）。
_PDF_CAL_WD = ["日", "月", "火", "水", "木", "金", "土"]


def _cal_col_index(d):
    """日曜始まりの列インデックス（Python weekday: 月=0..日=6 → 日=0..土=6）。"""
    return (d.weekday() + 1) % 7


def export_pdf_individual(
    shifts_data: list,
    staff_list: list,
    year: int,
    month: int,
    staff_ids: list = None,
    oncall_map: dict = None,
    cook_labels: dict = None,
) -> BytesIO:
    """職員1人につき1ページの月間カレンダーPDF（縦A4）。

    staff_ids: 対象 staff_id のリスト（None なら staff_list 全員）。staff_list は
    呼び出し側で在籍者のみに絞られている前提（休職者は含まれない）。
    """
    _register_cook_labels(cook_labels)
    dates, assignment_map, _summary_map, _phone_map, desk_slot_map, _break_map, bath_map, _meal_map = _build_daily_data(
        shifts_data, staff_list, year, month
    )
    parking_map = _build_parking_map(shifts_data)

    if staff_ids is not None:
        want = set(staff_ids)
        targets = [s for s in staff_list if s["id"] in want]
    else:
        targets = list(staff_list)

    # is_phone_duty を (staff_id → 日付集合) に集約（オンコール★表示用）
    phone_by_sid = {}
    for item in (shifts_data or []):
        if item.get("is_phone_duty"):
            phone_by_sid.setdefault(item["staff_id"], set()).add(item["date"])

    def cell_for(staff, d):
        """(表示テキスト, 勤務フラグ, オンコールフラグ) を返す。"""
        sid = staff["id"]
        is_cook = (staff.get("department") == "cooking")
        d_str = d.isoformat()
        if is_cook:
            asgn = assignment_map.get(d_str, {}).get(sid, "")
            text = ASSIGNMENT_LABELS.get(asgn, "")
        else:
            asgn, text = _care_cell_text(d_str, sid, assignment_map, bath_map, desk_slot_map)
        text = _append_parking(text, parking_map, d_str, sid)
        off_token = "cook_off" if is_cook else "off"
        working = asgn not in (off_token, "")
        if not text:
            # 休み（off）・未割当は「休」表示。空文字は当該日に行が無い＝休扱い。
            text = "休"
        is_oncall = d_str in phone_by_sid.get(sid, set())
        if oncall_map and oncall_map.get(d_str) == staff.get("name"):
            is_oncall = True
        return text, working, is_oncall

    return _render_pdf_individual_calendars(targets, dates, year, month, cell_for)


def _render_pdf_individual_calendars(targets, dates, year, month, cell_for):
    """縦A4・1人1ページの月間カレンダーを描画する。"""
    from fpdf import FPDF  # 遅延 import

    MARGIN = 10.0
    TITLE_H = 11.0
    WD_H = 7.0            # 曜日見出しの高さ
    GAP = 2.0
    FOOTER_H = 8.0

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(False)
    pdf.add_font(_PDF_FONT, "", _PDF_FONT_PATH)

    page_w = 210.0
    page_h = 297.0
    usable_w = page_w - MARGIN * 2
    col_w = usable_w / 7.0

    num_days = calendar.monthrange(year, month)[1]
    first = date(year, month, 1)
    start_col = _cal_col_index(first)
    n_weeks = (start_col + num_days + 6) // 7

    grid_top = MARGIN + TITLE_H + GAP + WD_H
    grid_bottom = page_h - MARGIN - FOOTER_H
    row_h = (grid_bottom - grid_top) / max(n_weeks, 1)

    day_num_fs = 10.0
    body_fs = 9.0

    def line_h(fs):
        return _pdf_line_h(fs)

    def text_block(x, y, w, h, text, fs, color=(0, 0, 0), align="C"):
        if not text:
            return
        max_lines = max(1, int(h / line_h(fs)))
        lines = _pdf_wrap(pdf, text, w, fs)[:max_lines]
        pdf.set_font(_PDF_FONT, "", fs)
        pdf.set_text_color(*color)
        lh = line_h(fs)
        ty = y + (h - lh * len(lines)) / 2
        for ln in lines:
            pdf.set_xy(x, ty)
            pdf.cell(w, lh, ln, align=align)
            ty += lh
        pdf.set_text_color(0, 0, 0)

    if not targets:
        # 空でも1ページは出す（「対象なし」）
        pdf.add_page()
        pdf.set_font(_PDF_FONT, "", 14)
        pdf.set_xy(MARGIN, MARGIN)
        pdf.cell(usable_w, TITLE_H, f"{year}年{month}月 個別シフト（対象職員なし）", align="C")
        buf = BytesIO()
        buf.write(bytes(pdf.output()))
        buf.seek(0)
        return buf

    for staff in targets:
        pdf.add_page()

        # タイトル
        pdf.set_font(_PDF_FONT, "", 15)
        pdf.set_xy(MARGIN, MARGIN)
        pdf.cell(usable_w, TITLE_H, f"{year}年{month}月 シフト表　{staff.get('name', '')}", align="C")

        # 曜日見出し（日=赤 / 土=青）
        wd_y = MARGIN + TITLE_H + GAP
        for c in range(7):
            x = MARGIN + col_w * c
            if c == 0:
                fill = _PDF_SUN_BG
            elif c == 6:
                fill = _PDF_SAT_BG
            else:
                fill = _PDF_SUMMARY_BG
            pdf.set_fill_color(*fill)
            pdf.rect(x, wd_y, col_w, WD_H, style="DF")
            col = (204, 0, 0) if c == 0 else ((0, 0, 204) if c == 6 else (0, 0, 0))
            text_block(x, wd_y, col_w, WD_H, _PDF_CAL_WD[c], 11.0, color=col)

        # まずカレンダー全マス（月外の空白マスも含む）に枠線を引く。
        # 日付のあるマスだけ描くと、月初・月末の空きマスに線が入らず欠けて見えるため。
        for wk in range(n_weeks):
            for col in range(7):
                pdf.rect(MARGIN + col_w * col, grid_top + row_h * wk, col_w, row_h, style="D")

        # 日付グリッド
        work_days = 0
        off_days = 0
        for day in range(1, num_days + 1):
            d = date(year, month, day)
            idx = start_col + (day - 1)
            wk = idx // 7
            col = idx % 7
            x = MARGIN + col_w * col
            y = grid_top + row_h * wk

            fill = _pdf_weekend_color(d)
            if fill:
                pdf.set_fill_color(*fill)
                pdf.rect(x, y, col_w, row_h, style="DF")
            else:
                pdf.rect(x, y, col_w, row_h, style="D")

            # 日番号（左上・日曜赤/土曜青/祝日赤）
            if jpholiday.is_holiday(d) or col == 0:
                dcol = (204, 0, 0)
            elif col == 6:
                dcol = (0, 0, 204)
            else:
                dcol = (0, 0, 0)
            pdf.set_font(_PDF_FONT, "", day_num_fs)
            pdf.set_text_color(*dcol)
            pdf.set_xy(x + 1.0, y + 0.8)
            pdf.cell(col_w - 2.0, line_h(day_num_fs), str(day), align="L")
            pdf.set_text_color(0, 0, 0)

            text, working, is_oncall = cell_for(staff, d)
            if working:
                work_days += 1
            else:
                off_days += 1
            body = ("★" + text) if is_oncall else text
            body_color = (204, 0, 0) if is_oncall else (0, 0, 0)
            # 日番号帯の下に本文
            top_pad = line_h(day_num_fs) + 1.2
            text_block(x, y + top_pad, col_w, row_h - top_pad, body, body_fs, color=body_color)

        # フッター（勤務日数）
        pdf.set_font(_PDF_FONT, "", 11.0)
        pdf.set_xy(MARGIN, grid_bottom + 1.5)
        pdf.cell(usable_w, FOOTER_H - 1.5,
                 f"勤務日数: {work_days}日　休み: {off_days}日　（★=オンコール）", align="R")

    buf = BytesIO()
    buf.write(bytes(pdf.output()))
    buf.seek(0)
    return buf


# 調理/介護サマリーのラベル（Excel解析でスタッフ行と区別する用）
_PDF_SUMMARY_LABELS = {
    "訪問午前", "訪問午後", "デイ午前", "デイ午後", "兼務者数", "オンコール", "調理配置数",
}


def export_pdf_from_excel(file_bytes, group: str = "care", half: str = "first") -> BytesIO:
    """アプリのExcel出力形式（手直し済み）から、該当PDFを描画する。
    保存データは一切使わない＝アップロードされたExcelのセル内容を忠実に描画。警告は再計算しない。
    """
    is_cook, group_label, _ = _pdf_group_meta(group)
    sheet_name = "調理スタッフ" if is_cook else "介護スタッフ"
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    # ヘッダ行（「職員名」セル）を特定
    header_row = name_col = None
    for r in range(1, min(ws.max_row, 8) + 1):
        for c in range(1, min(ws.max_column, 6) + 1):
            if str(ws.cell(row=r, column=c).value or "").strip() == "職員名":
                header_row, name_col = r, c
                break
        if header_row:
            break
    if header_row is None:
        raise ValueError("Excelの形式を認識できません（『職員名』ヘッダが見つかりません）。")

    # 日付列(M/D)と出勤日数列
    date_cols = []  # (col, month, day)
    total_col = None
    for c in range(name_col + 1, ws.max_column + 1):
        v = str(ws.cell(row=header_row, column=c).value or "").strip()
        mm = re.match(r"^(\d{1,2})/(\d{1,2})$", v)
        if mm:
            date_cols.append((c, int(mm.group(1)), int(mm.group(2))))
        elif "出勤" in v:
            total_col = c
    if not date_cols:
        raise ValueError("日付列(M/D)が見つかりません。アプリのExcel出力形式か確認してください。")

    title_text = str(ws.cell(row=1, column=1).value or "")
    ym = re.search(r"(\d{4})\s*年", title_text)
    year = int(ym.group(1)) if ym else date.today().year
    month = date_cols[0][1]

    def in_half(day):
        return (day <= 15) if half != "second" else (day >= 16)
    sel_cols = [t for t in date_cols if in_half(t[2])] or date_cols
    sel_dates = [date(year, mth, day) for (_c, mth, day) in sel_cols]

    staff_rows, summary_rows = [], []
    for r in range(header_row + 2, ws.max_row + 1):
        label = str(ws.cell(row=r, column=name_col).value or "").strip()
        if not label:
            continue
        cells = [str(ws.cell(row=r, column=c).value or "").strip() for (c, _m, _d) in sel_cols]
        if label in _PDF_SUMMARY_LABELS:
            summary_rows.append((label, cells, [False] * len(cells)))
        else:
            work = str(ws.cell(row=r, column=total_col).value or "").strip() if total_col else ""
            staff_rows.append((label, cells, work))

    half_label = "後半" if half == "second" else "前半"
    first_d = sel_cols[0][2]
    last_d = sel_cols[-1][2]
    title = f"{year}年{month}月 シフト表（{group_label} {half_label}：{first_d}〜{last_d}日）"
    buf = _render_pdf_table(title, sel_dates, staff_rows, summary_rows,
                            paper="A4" if group == "cooking" else "A3")
    # 依頼文26: ファイル名に対象月を入れるため year/month も返す
    return buf, year, month


# ===========================================================================
# 依頼文41: 手修正Excel → 保存シフトデータへの書き戻し（取り込み）
# ===========================================================================
# セル表示文字 → assignment コード 逆引き（現行コードを優先。旧名 day_am 等は出力しない）
_LABEL_TO_ASSIGNMENT = {
    "デイ8:30-17:30": "day_pattern1",
    "デイ9:00-16:00": "day_pattern2",
    "デイ午前のみ": "day_pattern3",
    "デイ午後のみ": "day_pattern4",
    "早番7:30-16:30": "early",
    "遅番9:30-18:30": "late",
    "看護9:30-13:30": "nurse_short",
    "訪問午前のみ": "visit_am",
    "訪問午後のみ": "visit_pm",
    "兼務(デイ→訪問)": "day_p3_visit_pm",
    "兼務(訪問→デイ)": "visit_am_day_p4",
    "調理①6-8": "cooking_1",
    "調理②8-13": "cooking_2",
    "調理③12-19": "cooking_3",
    "調理④6-13": "cooking_4",
    "調理⑤9-15": "cooking_5",
}


def parse_shift_cell(text):
    """セル表示文字 → {"assignment","bath_role","desk_slots"}。
    空セル＝{"assignment":"off",...}。解釈できないセルは None を返す（依頼文41・書き戻し不可）。
    """
    raw = (text or "").replace("\r", "").strip()
    if raw == "":
        return {"assignment": "off", "bath_role": None, "desk_slots": None}
    lines = [ln.strip() for ln in raw.split("\n") if ln.strip()]
    bath_role = None
    desk_slots = None
    base_lines = []
    for ln in lines:
        if ln in ("中介助", "外介助"):
            bath_role = ln[0]  # "中" / "外"
        elif ln == "コイン" or re.match(r"^P\w+$", ln):
            # 駐車場ラベル（依頼文24・表示専用）は割当ではないので無視する。
            # これを base_lines に入れると「基本ラベルが一意に取れない」で
            # 解釈不能になり、車通勤者のセルが全て書き戻し不可になってしまう。
            continue
        elif ln in ("相談（終日）", "相談(終日)"):
            desk_slots = [0, 1, 2, 3]
        elif ln.startswith("相談:") or ln.startswith("相談："):
            body = ln.split(":", 1)[-1] if ":" in ln else ln.split("：", 1)[-1]
            labels = [x.strip() for x in body.replace("、", ",").split(",") if x.strip()]
            idx = []
            for lab in labels:
                if lab in DESK_SLOT_LABELS:
                    idx.append(DESK_SLOT_LABELS.index(lab))
                else:
                    return None  # 未知の相談スロット表記＝解釈不能
            if not idx:
                return None
            desk_slots = sorted(set(idx))
        else:
            base_lines.append(ln)
    if len(base_lines) != 1:
        return None  # 基本ラベルが一意に取れない＝解釈不能
    assignment = _LABEL_TO_ASSIGNMENT.get(base_lines[0])
    if assignment is None:
        return None
    return {"assignment": assignment, "bath_role": bath_role, "desk_slots": desk_slots}


def state_to_cell_text(assignment, bath_role, desk_slots):
    """保存状態 → 表示文字（差分プレビュー用。export の _care_cell_text と同じ流儀）。"""
    if not assignment or assignment == "off":
        return "（休み）"
    text = ASSIGNMENT_LABELS.get(assignment, assignment)
    if bath_role:
        text += f" {bath_role}介助"
    if desk_slots:
        if set(desk_slots) >= {0, 1, 2, 3}:
            text += " 相談（終日）"
        else:
            labs = [DESK_SLOT_LABELS[i] for i in desk_slots if i < len(DESK_SLOT_LABELS)]
            if labs:
                text += " 相談:" + ",".join(labs)
    return text


def parse_uploaded_shift_excel(file_bytes, group="care", half="first"):
    """依頼文41: 手修正Excelを読み、職員名×日付の保存状態(逆引き)を返す。

    Returns dict:
      year, month, group, half, is_cook,
      date_isos: [iso...]（対象範囲の日付）,
      staff_names: [name...]（行順・資格表記は除いた氏名）,
      cells: {name: {iso: state_or_None}}（None=解釈不能セル）,
      unparseable: [{"name","date","text"}...],
    raise ValueError 形式不一致。
    """
    is_cook = (group == "cooking")
    sheet_name = "調理スタッフ" if is_cook else "介護スタッフ"
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    header_row = name_col = None
    for r in range(1, min(ws.max_row, 8) + 1):
        for c in range(1, min(ws.max_column, 6) + 1):
            if str(ws.cell(row=r, column=c).value or "").strip() == "職員名":
                header_row, name_col = r, c
                break
        if header_row:
            break
    if header_row is None:
        raise ValueError("Excelの形式を認識できません（『職員名』ヘッダが見つかりません）。")

    date_cols = []
    for c in range(name_col + 1, ws.max_column + 1):
        v = str(ws.cell(row=header_row, column=c).value or "").strip()
        mm = re.match(r"^(\d{1,2})/(\d{1,2})$", v)
        if mm:
            date_cols.append((c, int(mm.group(1)), int(mm.group(2))))
    if not date_cols:
        raise ValueError("日付列(M/D)が見つかりません。アプリのExcel出力形式か確認してください。")

    title_text = str(ws.cell(row=1, column=1).value or "")
    ym = re.search(r"(\d{4})\s*年", title_text)
    year = int(ym.group(1)) if ym else date.today().year
    month = date_cols[0][1]

    def in_half(day):
        return (day <= 15) if half != "second" else (day >= 16)
    sel_cols = [t for t in date_cols if in_half(t[2])] or date_cols
    date_isos = [date(year, m, d).isoformat() for (_c, m, d) in sel_cols]

    cells = {}
    staff_names = []
    unparseable = []
    for r in range(header_row + 2, ws.max_row + 1):
        label = str(ws.cell(row=r, column=name_col).value or "").strip()
        if not label:
            continue
        name = label.split("\n")[0].strip()
        if label in _PDF_SUMMARY_LABELS or name in _PDF_SUMMARY_LABELS:
            continue  # サマリー行はスキップ
        row_states = {}
        for (c, _m, _d), iso in zip(sel_cols, date_isos):
            raw = str(ws.cell(row=r, column=c).value or "").strip()
            st = parse_shift_cell(raw)
            row_states[iso] = st
            if st is None:
                unparseable.append({"name": name, "date": iso, "text": raw})
        cells[name] = row_states
        staff_names.append(name)

    return {
        "year": year, "month": month, "group": group, "half": half, "is_cook": is_cook,
        "date_isos": date_isos, "staff_names": staff_names,
        "cells": cells, "unparseable": unparseable,
    }


def recompute_warnings_from_shifts(shifts_data, staff_list, settings, year, month):
    """依頼文41-D: 保存シフト(手修正後)に対し、主要な条件未達を再計算して警告を返す。
    ソルバーは使わず、最低人数系の代表的な制約のみを後処理で検査する（手修正優先＝拒否しない）。
    検査対象: デイ午前/午後 最低・訪問午前/午後 最低・中介助/外介助 最低・早番/遅番 各1名・相談員desk 1名。
    """
    warnings = []
    num_days = calendar.monthrange(year, month)[1]
    closed = set(settings.get("closed_days", []) or [])
    visit_days = set(settings.get("visit_operating_days", []) or [])
    min_day = int(settings.get("min_day_service", 0) or 0)
    min_vam = int(settings.get("min_visit_am", 0) or 0)
    min_vpm = int(settings.get("min_visit_pm", 0) or 0)
    min_mid = int(settings.get("min_bath_mid", 0) or 0)
    min_out = int(settings.get("min_bath_out", 0) or 0)
    min_early = int(settings.get("min_early_staff", 1) or 0)
    min_late = int(settings.get("min_late_staff", 1) or 0)
    # デイ利用者がいない曜日は介護を原則N名（既定2名）に抑える（デイ人数の下限は課さない）
    no_ds_days = set(settings.get("no_day_service_days", []) or [])
    no_ds_min = int(settings.get("no_day_service_min_staff", 2) or 0)

    nurse_pt_ids = {st["id"] for st in staff_list if _is_nurse_or_pt_staff(st)}
    counselor_qual_ids = _get_counselor_qual_ids_for_validation(settings)

    by_date = {}
    for it in shifts_data:
        by_date.setdefault(it["date"], []).append(it)

    for d in range(1, num_days + 1):
        dt = date(year, month, d)
        if dt.weekday() in closed:
            continue
        iso = dt.isoformat()
        items = by_date.get(iso, [])
        day_am = sum(1 for it in items if it["assignment"] in _DAY_AM_SET and it["staff_id"] not in nurse_pt_ids)
        day_pm = sum(1 for it in items if it["assignment"] in _DAY_PM_SET and it["staff_id"] not in nurse_pt_ids)
        is_visit_day = dt.weekday() in visit_days
        v_am = sum(1 for it in items if it["assignment"] in _VISIT_AM_SET)
        if is_visit_day:
            # 訪問営業日の早番は「午前訪問＋午後デイ」＝訪問午前の担い手として数える
            v_am += sum(1 for it in items if it["assignment"] == "early")
        v_pm = sum(1 for it in items if it["assignment"] in _VISIT_PM_SET)
        n_mid = sum(1 for it in items if it.get("bath_role") == "中")
        n_out = sum(1 for it in items if it.get("bath_role") == "外")
        n_early = sum(1 for it in items if it["assignment"] == "early")
        n_late = sum(1 for it in items if it["assignment"] == "late")
        n_desk = sum(1 for it in items if it.get("counselor_desk_slots"))

        def warn(wt, msg):
            warnings.append({"date": iso, "warning_type": wt, "message": msg})

        if dt.weekday() in no_ds_days:
            # デイ利用者がいない曜日はデイ人数の下限を課さず、介護人数の超過だけ見る
            if no_ds_min > 0:
                care_work = sum(
                    1 for it in items
                    if it["staff_id"] not in nurse_pt_ids
                    and it["assignment"] in _CARE_WORK_SET
                )
                if care_work > no_ds_min:
                    warn(
                        "over_staffed_no_day_service",
                        f"デイ以外の曜日: 介護{care_work}名"
                        f"（原則{no_ds_min}名／{care_work - no_ds_min}名超過）",
                    )
        else:
            if min_day > 0 and day_am < min_day:
                warn("understaffed_day_am", f"デイサービス午前: {min_day - day_am}名不足")
            if min_day > 0 and day_pm < min_day:
                warn("understaffed_day_pm", f"デイサービス午後: {min_day - day_pm}名不足")
        if is_visit_day:
            if min_vam > 0 and v_am < min_vam:
                warn("understaffed_visit_am", f"訪問介護午前: {min_vam - v_am}名不足")
            if min_vpm > 0 and v_pm < min_vpm:
                warn("understaffed_visit_pm", f"訪問介護午後: {min_vpm - v_pm}名不足")
        if min_mid > 0 and n_mid < min_mid:
            warn("bath_mid_short", f"中介助 {min_mid - n_mid}名不足（必要{min_mid}名・配置{n_mid}名）")
        if min_out > 0 and n_out < min_out:
            warn("bath_out_short", f"外介助 {min_out - n_out}名不足（必要{min_out}名・配置{n_out}名）")
        if min_early > 0 and n_early < min_early:
            warn("early_unassigned", f"早番(7:30-16:30): {min_early - n_early}名不足（早番未配置）")
        if min_late > 0 and n_late < min_late:
            warn("late_unassigned", f"遅番(9:30-18:30): {min_late - n_late}名不足（遅番未配置）")
        if counselor_qual_ids and n_desk == 0:
            warn("counselor_unassigned", "相談員 未配置")

    return warnings


def _get_counselor_qual_ids_for_validation(settings):
    """配置ルールから相談員資格IDを拾う（warning再計算用・無ければ空集合）。"""
    ids = set()
    for pr in settings.get("placement_rules", []) or []:
        nm = pr.get("name", "")
        if "相談" in nm or pr.get("rule_type") == "qualification_min" and "相談" in nm:
            ids.update(pr.get("target_qualification_ids", []) or [])
    return ids


def export_excel_group_half(
    shifts_data: list,
    warnings_data: list,
    staff_list: list,
    year: int,
    month: int,
    group: str = "care",
    half: str = "first",
    oncall_map: dict = None,
    cook_labels: dict = None,
) -> BytesIO:
    """Excelを「PDFと同じ4分割」で出力（1シート）。データ・並びは既存Excelと同じ。"""
    _register_cook_labels(cook_labels)
    dates, assignment_map, summary_map, phone_duty_map, desk_slot_map, break_map, bath_map, meal_map = _build_daily_data(
        shifts_data, staff_list, year, month
    )
    if oncall_map is not None:
        phone_duty_map = {d: [n] for d, n in oncall_map.items() if n}

    is_cook = (group == "cooking")
    gstaff = [s for s in staff_list if (s.get("department") == "cooking") == is_cook]
    sel, half_label = _half_dates(dates, half)

    # PDFと同じタイトル書式（例:「2026年7月 シフト表（調理 前半：1〜15日）」）
    group_label = "調理" if is_cook else "介護・看護"
    first_d = sel[0].day if sel else 1
    last_d = sel[-1].day if sel else 1
    title = f"{year}年{month}月 シフト表（{group_label} {half_label}：{first_d}〜{last_d}日）"

    wb = Workbook()
    ws = wb.active
    ws.title = "調理スタッフ" if is_cook else "介護スタッフ"
    _write_group_sheet(
        ws, gstaff, sel, year, month,
        assignment_map=assignment_map, summary_map=summary_map,
        phone_duty_map=phone_duty_map, desk_slot_map=desk_slot_map,
        bath_map=bath_map, warnings_data=warnings_data, is_cook=is_cook,
        fit_one_page=True, parking_map=_build_parking_map(shifts_data),
        title_override=title,
    )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
